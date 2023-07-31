import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import load_workbook

def create_excel_sheet(file_path, sheet_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    return workbook, sheet

def load_existing_sheet(file_path, sheet_name):
    workbook = load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
        sheet.title = sheet_name
    return workbook, sheet

def add_data_to_sheet(sheet, roll_number, name):
    sheet.append([roll_number, name])

def save_workbook(workbook, file_path):
    workbook.save(file_path)

def save_data():
    roll_number = roll_entry.get()
    name = name_entry.get()

    if not roll_number or not name:
        messagebox.showwarning("Warning", "Please enter both Roll Number and Name.")
        return

    try:
        add_data_to_sheet(sheet, roll_number, name)
        save_workbook(workbook, excel_file_path)
        messagebox.showinfo("Success", "Data saved successfully in 'data.xlsx'")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

    # Clear the input fields after saving data
    roll_entry.delete(0, tk.END)
    name_entry.delete(0, tk.END)

if __name__ == "__main__":
    excel_file_path = "C:\\Users\\Surya\\Documents\\data.xlsx"
    sheet_name = "StudentData"

    try:
        # Check if the file and sheet already exist, load them if they do
        try:
            workbook, sheet = load_existing_sheet(excel_file_path, sheet_name)
        except FileNotFoundError:
            workbook, sheet = create_excel_sheet(excel_file_path, sheet_name)

        root = tk.Tk()
        root.title("Student Data Entry")
        root.geometry("300x150")

        roll_label = tk.Label(root, text="Roll Number:")
        roll_label.pack()
        roll_entry = tk.Entry(root)
        roll_entry.pack()

        name_label = tk.Label(root, text="Name:")
        name_label.pack()
        name_entry = tk.Entry(root)
        name_entry.pack()

        save_button = tk.Button(root, text="Save Data", command=save_data)
        save_button.pack()

        root.mainloop()
    
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")
